"""Write normalized rows to dated CSV and XLSX files under data/."""
from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

import config
from parse import ROW_COLUMNS

# Columns shown in the formatted XLSX Products sheet (the *_json blobs are dropped
# there for readability; the CSV keeps everything).
XLSX_COLUMNS = [c for c in ROW_COLUMNS if not c.endswith("_json")]
# Numeric columns get real number types in XLSX.
_NUMERIC = {"unit_price", "quantity_available", "min_order_qty"}


def output_path(scrape_date: str, data_dir: Path = config.DATA_DIR) -> Path:
    return data_dir / f"excelta_{scrape_date}.csv"


def xlsx_path(scrape_date: str, data_dir: Path = config.DATA_DIR) -> Path:
    return data_dir / f"excelta_{scrape_date}.xlsx"


def write_csv(rows: list[dict], scrape_date: str, data_dir: Path = config.DATA_DIR) -> Path:
    data_dir.mkdir(parents=True, exist_ok=True)
    path = output_path(scrape_date, data_dir)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ROW_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return path


def _to_number(val):
    try:
        f = float(val)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return val


def write_xlsx(rows: list[dict], scrape_date: str, data_dir: Path = config.DATA_DIR) -> Path:
    """Formatted workbook: Products sheet (frozen header + autofilter) + Summary sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data_dir.mkdir(parents=True, exist_ok=True)
    path = xlsx_path(scrape_date, data_dir)

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="C00000")  # DigiKey red

    ws.append(XLSX_COLUMNS)
    for ci, _ in enumerate(XLSX_COLUMNS, start=1):
        c = ws.cell(row=1, column=ci)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")

    for r in rows:
        ws.append([
            _to_number(r.get(col, "")) if col in _NUMERIC else r.get(col, "")
            for col in XLSX_COLUMNS
        ])

    # Currency format for unit_price.
    price_idx = XLSX_COLUMNS.index("unit_price") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=price_idx).number_format = "$#,##0.000"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(XLSX_COLUMNS))}{ws.max_row}"

    # Column widths (capped).
    widths = {"description": 42, "detailed_description": 48, "datasheet_url": 40,
              "product_url": 40, "photo_url": 40, "category_name": 26}
    for ci, col in enumerate(XLSX_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 18)

    # --- Summary sheet ---
    sm = wb.create_sheet("Summary")
    in_stock = sum(1 for r in rows if _to_number(r.get("quantity_available")) not in ("", 0))
    total_qty = sum(v for r in rows if isinstance(v := _to_number(r.get("quantity_available")), int))
    sm.append(["Excelta @ DigiKey — Inventory Summary"])
    sm["A1"].font = Font(bold=True, size=14)
    sm.append([])
    for label, val in [
        ("Scrape date", scrape_date),
        ("Total parts", len(rows)),
        ("Parts in stock (qty > 0)", in_stock),
        ("Total units available", total_qty),
        ("Categories", len({r.get("category_name") for r in rows})),
    ]:
        sm.append([label, val])
        sm.cell(row=sm.max_row, column=1).font = Font(bold=True)

    sm.append([])
    sm.append(["Category", "Parts"])
    hdr = sm.max_row
    for col in (1, 2):
        sm.cell(row=hdr, column=col).font = header_font
        sm.cell(row=hdr, column=col).fill = header_fill
    for cat, n in Counter(r.get("category_name", "") for r in rows).most_common():
        sm.append([cat, n])
    sm.column_dimensions["A"].width = 34
    sm.column_dimensions["B"].width = 14

    wb.save(path)
    return path
