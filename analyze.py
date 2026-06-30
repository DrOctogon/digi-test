#!/usr/bin/env python3
"""Build the baseline analysis deliverables from the latest snapshot CSV.

  python analyze.py            # uses latest data/excelta_*.csv
  python analyze.py --date 2026-06-30

Outputs:
  data/excelta_analysis_<date>.xlsx   (Exec Summary + 5 analysis sheets)
  data/ANALYSIS_<date>.md             (Pyramid-Principle exec summary)
"""
from __future__ import annotations

import argparse
from pathlib import Path

import analysis as A
import config

RED = "C00000"


def _hdr(ws, row_idx, ncols):
    from openpyxl.styles import Font, PatternFill
    for ci in range(1, ncols + 1):
        c = ws.cell(row=row_idx, column=ci)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=RED)


def _table(ws, columns, rows, start_row=1):
    from openpyxl.utils import get_column_letter
    ws.append(columns)
    _hdr(ws, start_row, len(columns))
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    for ci, c in enumerate(columns, start=1):
        width = max(12, min(48, max((len(str(r.get(c, ""))) for r in rows[:200]), default=10) + 2))
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = f"A{start_row + 1}"


def build_xlsx(res: dict, date: str, total_skus: int, data_dir: Path = config.DATA_DIR) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    path = data_dir / f"excelta_analysis_{date}.xlsx"
    wb = Workbook()

    # --- Exec Summary ---
    ws = wb.active
    ws.title = "Exec Summary"
    abc, stock, price, assort, fp = (res["abc"], res["stock"], res["price"],
                                     res["assortment"], res["footprint"])
    A_ = abc["tier_summary"]["A"]
    lines = [
        ("Excelta @ DigiKey — Baseline Analysis (T=0)", ""),
        ("Scrape date", date),
        ("Data quality", res["quality"].notes[0]),
        ("", ""),
        ("PORTFOLIO", ""),
        ("Total SKUs", total_skus),
        ("Valued SKUs (price>0 & stock>0)", abc["valued_sku_count"]),
        ("Total inventory value (proxy)", f"${abc['total_inventory_value']:,.0f}"),
        ("A-tier: SKUs / % of SKUs / % of value",
         f"{A_['skus']} / {A_['skus']/max(abc['valued_sku_count'],1):.0%} / {A_['value_share']:.0%}"),
        ("", ""),
        ("AVAILABILITY", ""),
        ("Overall stockout rate", f"{stock['overall_stockout_rate']:.0%}"),
        ("Units available", f"{stock['total_units']:,}"),
        ("Deepest category", f"{fp['deepest_category']} ({fp['deepest_category_skus']} SKUs)"),
        ("", ""),
        ("PRICING", ""),
        ("Median list price (price>0)", f"${price['price_median']:,.2f}"),
        ("P90 / Max", f"${price['price_p90']:,.0f} / ${price['price_max']:,.0f}"),
        ("Quote-only ($0) SKUs", price["zero_price_count"]),
        ("SKUs w/ volume discount", price["skus_with_volume_discount"]),
        ("", ""),
        ("ASSORTMENT", ""),
        ("Rationalization candidates (dead/quote)", assort["rationalization_candidates"]),
        ("", ""),
        ("GUARDRAIL", res["quality"].notes[1]),
        ("NOTE", "Single snapshot = baseline only. Trend/demand needs >=2 daily snapshots."),
    ]
    for a, b in lines:
        ws.append([a, b])
    ws["A1"].font = Font(bold=True, size=14)
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == "" and ws.cell(row=r, column=1).value:
            ws.cell(row=r, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 40

    # --- Data Quality ---
    wq = wb.create_sheet("Data Quality")
    q = res["quality"]
    _table(wq, ["check", "count"],
           [{"check": "TOTAL ROWS", "count": q.total}] +
           [{"check": k, "count": v} for k, v in q.issues.items()])

    # --- ABC (top 200 + tier rollup) ---
    wa = wb.create_sheet("ABC")
    abc_cols = ["abc_tier", "digikey_part_number", "mfr_part_number", "category_name",
                "description", "unit_price", "quantity_available", "inventory_value",
                "cum_value_share"]
    _table(wa, abc_cols, abc["rows"][:200])
    wa.append([])
    wa.append(["TIER", "SKUs", "VALUE", "VALUE_SHARE"])
    _hdr(wa, wa.max_row, 4)
    for t in ("A", "B", "C"):
        s = abc["tier_summary"][t]
        wa.append([t, s["skus"], s["value"], s["value_share"]])

    # --- Stock Health ---
    wsh = wb.create_sheet("Stock Health")
    _table(wsh, ["category_name", "parts", "in_stock", "stockout_rate", "units_available",
                 "inventory_value", "oos_list_exposure"], stock["categories"])

    # --- Price Architecture ---
    wp = wb.create_sheet("Price Architecture")
    _table(wp, ["category_name", "priced_skus", "median_price", "min_price", "max_price",
                "price_spread_ratio", "coef_of_variation"], price["category_dispersion"])

    # --- Assortment ---
    was = wb.create_sheet("Assortment")
    _table(was, ["bucket", "skus"],
           [{"bucket": k, "skus": v} for k, v in assort["buckets"].items()])

    _add_charts(wb, abc, stock)
    wb.save(path)
    return path


def _add_charts(wb, abc, stock) -> None:
    """Pareto curve (ABC) + inventory-value and stockout-rate bars (Stock Health)."""
    from openpyxl.chart import BarChart, LineChart, Reference

    # Pareto: cumulative value share vs SKU rank (ABC sheet, col 9 = cum_value_share).
    wa = wb["ABC"]
    n = min(200, abc["valued_sku_count"])
    if n >= 2:
        pareto = LineChart()
        pareto.title = "Pareto — cumulative value share by SKU rank"
        pareto.y_axis.title = "cumulative value share"
        pareto.x_axis.title = "SKU rank (by inventory value)"
        data = Reference(wa, min_col=9, min_row=1, max_row=1 + n)
        pareto.add_data(data, titles_from_data=True)
        pareto.height, pareto.width = 9, 18
        wa.add_chart(pareto, "K2")

    # Stock Health bars: inventory_value (col 6) and stockout_rate (col 4) by category.
    wsh = wb["Stock Health"]
    ncat = len(stock["categories"])
    if ncat >= 2:
        last = 1 + min(15, ncat)  # top categories by value (already sorted)
        cats = Reference(wsh, min_col=1, min_row=2, max_row=last)

        bar_val = BarChart()
        bar_val.title = "Inventory value by category (proxy)"
        bar_val.add_data(Reference(wsh, min_col=6, min_row=1, max_row=last), titles_from_data=True)
        bar_val.set_categories(cats)
        bar_val.height, bar_val.width = 9, 18
        bar_val.legend = None
        wsh.add_chart(bar_val, "I2")

        bar_oos = BarChart()
        bar_oos.title = "Stockout rate by category"
        bar_oos.add_data(Reference(wsh, min_col=4, min_row=1, max_row=last), titles_from_data=True)
        bar_oos.set_categories(cats)
        bar_oos.height, bar_oos.width = 9, 18
        bar_oos.legend = None
        wsh.add_chart(bar_oos, "I20")


def build_markdown(res: dict, date: str, total_skus: int, data_dir: Path = config.DATA_DIR) -> Path:
    abc, stock, price, assort, fp = (res["abc"], res["stock"], res["price"],
                                     res["assortment"], res["footprint"])
    A_ = abc["tier_summary"]["A"]
    top = sorted(stock["categories"], key=lambda c: c["inventory_value"], reverse=True)[:5]
    worst = sorted([c for c in stock["categories"] if c["parts"] >= 20],
                   key=lambda c: c["stockout_rate"], reverse=True)[:5]
    path = data_dir / f"ANALYSIS_{date}.md"
    md = f"""# Excelta @ DigiKey — Baseline Analysis ({date})

> Data quality: **{res['quality'].notes[0]}** · {total_skus} SKUs · single snapshot (**T=0**).
> {A.INV_VALUE_PROXY_NOTE}. Trend/demand analysis is **gated** until >=2 daily snapshots exist.

## Bottom line (answer-first)
1. **Value is extreme-Pareto.** {A_['skus']} SKUs ({A_['skus']/max(abc['valued_sku_count'],1):.0%} of stocked SKUs) hold **{A_['value_share']:.0%}** of the **${abc['total_inventory_value']:,.0f}** visible inventory value. Manage these as A-class.
2. **Availability is the biggest leak.** Overall stockout rate **{stock['overall_stockout_rate']:.0%}** — over half the catalog is unbuyable right now, concentrated in high-value lines.
3. **Pricing is wide and under-laddered.** Median list **${price['price_median']:,.2f}**, max **${price['price_max']:,.0f}**; only **{price['skus_with_volume_discount']}** SKUs expose a volume discount.

## Lens 1 — Distributor / category management
Inventory value by top category (proxy):

| Category | Parts | In stock | Stockout | Inv. value |
|---|--:|--:|--:|--:|
""" + "\n".join(
        f"| {c['category_name']} | {c['parts']} | {c['in_stock']} | {c['stockout_rate']:.0%} | ${c['inventory_value']:,.0f} |"
        for c in top
    ) + f"""

Highest-stockout sizable categories (availability risk):

| Category | Parts | Stockout rate |
|---|--:|--:|
""" + "\n".join(
        f"| {c['category_name']} | {c['parts']} | {c['stockout_rate']:.0%} |" for c in worst
    ) + f"""

**Move:** prioritize replenishment/sourcing where stockout rate AND inventory value are both high; rationalize the **{assort['rationalization_candidates']}** dead/quote-only SKUs.

## Lens 2 — Pricing strategy
- Priced SKUs: **{price['priced_sku_count']}**; quote-only ($0): **{price['zero_price_count']}**.
- Distribution (right-skewed): median **${price['price_median']:,.2f}**, P90 **${price['price_p90']:,.0f}**, max **${price['price_max']:,.0f}**.
- Volume-discount coverage: **{price['skus_with_volume_discount']}** SKUs; median discount depth **{(price['median_discount_depth'] or 0):.0%}**.
- Highest intra-category price dispersion (peer-group inconsistency to investigate):
""" + "\n".join(
        f"  - {c['category_name']}: {c['priced_skus']} SKUs, spread {c['price_spread_ratio']}x, CV {c['coef_of_variation']}"
        for c in price["category_dispersion"][:5]
    ) + f"""

**Move:** standardize price ladders on A-tier SKUs; resolve $0/quote-only listings (lost conversion).

## Lens 3 — Competitive / market footprint
- Catalog footprint: **{fp['categories_covered']}** categories, **{fp['total_skus']}** SKUs; deepest = **{fp['deepest_category']}** ({fp['deepest_category_skus']} SKUs).
- Availability competitiveness (in-stock share): **{fp['availability_competitiveness']:.0%}**.
- *Caveat:* {fp['breadth_note']}. True competitive benchmarking needs peer-manufacturer snapshots (a future scraper extension).

## What we deliberately did NOT claim (no mistakes)
- No trend, churn, restock-velocity, or demand numbers — impossible from one snapshot.
- Inventory value is a **list x visible-stock proxy**, not COGS or sell-through.
- $0-price SKUs excluded from price stats; qty=0 counted as a real state, never imputed.

## Next
- Accumulate daily snapshots -> unlock WS6 (restock velocity, stockout duration, price-change detection, new/discontinued SKU flow).
"""
    path.write_text(md, encoding="utf-8")
    return path


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Baseline analysis of the latest Excelta snapshot")
    ap.add_argument("--date", default=None, help="Snapshot date YYYY-MM-DD (default: latest)")
    args = ap.parse_args(argv)

    if args.date:
        csv_path = config.DATA_DIR / f"excelta_{args.date}.csv"
    else:
        csv_path = A.latest_csv()
    date = csv_path.stem.replace("excelta_", "")

    rows = A.load_rows(csv_path)
    res = A.run_all(rows)
    xlsx = build_xlsx(res, date, total_skus=len(rows))
    md = build_markdown(res, date, total_skus=len(rows))
    print(f"Analyzed {len(rows)} SKUs from {csv_path.name}")
    print(f"  -> {xlsx}")
    print(f"  -> {md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
